import csv
import openpyxl
import pickle
from copy import copy
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, Union

import pandas as pd
import numpy as np

__all__ = ("save_to_csv", "read_from_pickle", "save_to_pickle")

DATE = date.today().strftime("%d-%m-%y")

# -----------------------------------------------------------
# csv
# -----------------------------------------------------------


def save_to_csv(data, filename: str, columns: list):
    filename = f"{filename}_{DATE}.csv"
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(data)


# -----------------------------------------------------------
# excel
# -----------------------------------------------------------


def _get_column_width(df: pd.DataFrame, col_number: int, max_col_width: int = 30) -> int:
    width = max(
        df.iloc[:, col_number].astype(str).str.len().max(),
        len(df.columns[col_number]) + 6,
    )
    return min(max_col_width, width)


def _format_data_in_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    df: pd.DataFrame,
    first_col: int,
    fmt_int: str = "#.##0",
    fmt_float: str = "#.##0,00",
) -> openpyxl.worksheet.worksheet.Worksheet:
    def set_column_format(ws, column_letter, fmt):
        for cell in ws[column_letter]:
            cell.number_format = fmt

    for col_index, data_type in enumerate(df.dtypes, first_col):
        col_number = col_index - first_col
        width = _get_column_width(df=df, col_number=col_number)
        column_letter = get_column_letter(col_index)
        worksheet.column_dimensions[column_letter].width = width
        if np.issubdtype(data_type, np.integer):
            set_column_format(worksheet, column_letter, fmt_int)
        if np.issubdtype(data_type, np.floating):
            set_column_format(worksheet, column_letter, fmt_float)
    return worksheet


def _copy_excel_cell_range(
    src_ws: openpyxl.worksheet.worksheet.Worksheet,
    min_row: int = None,
    max_row: int = None,
    min_col: int = None,
    max_col: int = None,
    tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
    tgt_min_row: int = 1,
    tgt_min_col: int = 1,
    with_style: bool = True,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    copies all cells from the source worksheet [src_ws] starting from [min_row] row
    and [min_col] column up to [max_row] row and [max_col] column
    to target worksheet [tgt_ws] starting from [tgt_min_row] row
    and [tgt_min_col] column.
    """
    if tgt_ws is None:
        tgt_ws = src_ws

    for row in src_ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if type(cell).__name__ == "MergedCell":
                continue
            tgt_cell = tgt_ws.cell(
                row=cell.row + tgt_min_row - min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value,
            )
            if with_style and cell.has_style:
                # tgt_cell._style = copy(cell._style)
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws


def append_df_to_excel(
    filename: Union[str, Path],
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    startrow: Optional[int] = None,
    min_row: Optional[int] = None,
    autofilter: bool = False,
    fmt_date: str = "dd-mm-yyyy",
    fmt_datetime: str = "dd-mm-yyyy hh:mm",
    truncate_sheet: bool = False,
    storage_options: Optional[dict] = None,
    **to_excel_kwargs,
) -> None:
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet. If [filename] doesn't exist, then this function will create it.

    storage_options: dict, optional
        Extra options that make sense for a particular storage connection, e.g. host, port,
        username, password, etc., if using a URL that will be parsed by fsspec, e.g.,
        starting “s3://”, “gcs://”.
    to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]

    Usage examples:

    >>> append_df_to_excel('/tmp/test.xlsx', df, autofilter=True,
                           freeze_panes=(1,0))

    >>> append_df_to_excel('/tmp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    >>> append_df_to_excel('/tmp/test.xlsx', df, index=False,
                           fmt_datetime="dd.mm.yyyy hh:mm")

    """
    filename = Path(filename)
    file_exists = filename.is_file()

    # if df is written using `index=True`, then `first_col = 2`, else `first_col = 1`
    first_col = int(to_excel_kwargs.get("index", True)) + 1
    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    # save content of existing sheets
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(
        filename,
        # filename.with_suffix(".xlsx"),
        engine="openpyxl",
        mode="a" if file_exists else "w",
        if_sheet_exists="new" if file_exists else None,
        date_format=fmt_date,
        datetime_format=fmt_datetime,
        storage_options=storage_options,
    ) as writer:
        if file_exists:
            # open an existing workbook
            writer.book = wb

            # get the last row in the existing Excel sheet if not specified
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet -> replace specified sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])
                writer.book.create_sheet(sheet_name, idx)
            # copy existing sheets
            writer.sheets = sheets
        else:
            # file doesn't exist -> create new one
            startrow = 0

        # write out the dataframe to excelwriter
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        worksheet = writer.sheets[sheet_name]

        if autofilter:
            worksheet.auto_filter.ref = worksheet.dimensions

        worksheet = _format_data_in_sheet(worksheet=worksheet, df=df, first_col=first_col)

    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)
        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]
        # copy rows written by `df.to_excel(...)` to
        _copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
            tgt_min_row=startrow,
            with_style=True,
            min_row=min_row,
        )
        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()


# -----------------------------------------------------------
# pickle
# -----------------------------------------------------------


def read_from_pickle(filename: str):
    with open(filename, "rb") as f:
        data = pickle.load(f)
    return data


def save_to_pickle(data, filename: str) -> None:
    with open(f"{filename}_{DATE}.pickle", "wb") as f:
        pickle.dump(data, f, protocol=pickle.HIGHEST_PROTOCOL)
