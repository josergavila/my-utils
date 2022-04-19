import csv
import openpyxl
import pickle
from copy import copy
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, Union

import pandas as pd
import numpy as np

from file_utils import check_if_file_exists


def _copy_excel_cell_range(
    src_ws: openpyxl.worksheet.worksheet.Worksheet,
    min_row: int = None,
    max_row: int = None,
    min_col: int = None,
    max_col: int = None,
    tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
    tgt_min_row: int = 1,
    tgt_min_col: int = 1,
    with_style: bool = True,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    copies all cells from the source worksheet [src_ws] starting from [min_row] row
    and [min_col] column up to [max_row] row and [max_col] column
    to target worksheet [tgt_ws] starting from [tgt_min_row] row
    and [tgt_min_col] column.
    """
    if tgt_ws is None:
        tgt_ws = src_ws

    for row in src_ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if type(cell).__name__ == "MergedCell":
                continue
            tgt_cell = tgt_ws.cell(
                row=cell.row + tgt_min_row - min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value,
            )
            if with_style and cell.has_style:
                # tgt_cell._style = copy(cell._style)
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws


def get_sheets_in_workbook(wb: openpyxl.workbook) -> dict:
    return {ws.title: ws for ws in wb.worksheets}


def get_max_row(ws: openpyxl.worksheet.worksheet) -> int:
    """calculates last row in worksheet with data (only works for contiguous data)"""
    max_row = ws.max_row
    for row in reversed(list(ws)):
        if not all(cell.value is None for cell in row):
            break
        max_row -= 1
    return max_row


def clear_sheet(wb: openpyxl.workbook, sheet_name: str) -> openpyxl.workbook:
    """clear sheet values (delete and create a new sheet)"""
    idx = wb.sheetnames.index(sheet_name)
    wb.remove(wb.worksheets[idx])
    wb.create_sheet(sheet_name, idx)
    return wb


def get_start_col(ws: openpyxl.worksheet.worksheet) -> int:
    """calculates first col in worksheet with data"""
    for start_col, col in enumerate(ws.iter_cols(), 1):
        if not all(cell.value is None for cell in col):
            return start_col
    return 1


def format_data_in_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    df: pd.DataFrame,
    index: bool = True,
    fmt_int: str = "#,##0",
    fmt_float: str = "#,##0.00",
) -> openpyxl.worksheet.worksheet.Worksheet:
    for i, data_type in enumerate(df.dtypes, 1):
        col_letter = get_column_letter(i + int(index))
        width = get_column_width(df=df, col_number=i)
        ws = set_column_width(ws, col_letter, width=width)
        if data_type in [np.dtype(np.int_), np.floating]:
            fmt = fmt_float if np.issubdtype(data_type, np.floating) else fmt_int
            ws = set_number_format(ws, col_letter, fmt)
    return ws


def get_column_width(df: pd.DataFrame, col_number: int, max_col_width: int = 30) -> int:
    width = max(
        df.iloc[:, col_number - 1].astype(str).str.len().max(),
        len(str(df.columns[col_number - 1])) + 6,
    )
    return min(max_col_width, width)


def set_column_width(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    col_letter: str,
    width: int = 6,
):
    ws.column_dimensions[col_letter].width = width
    return ws


def set_number_format(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    col_letter: str,
    fmt: str,
) -> openpyxl.worksheet.worksheet.Worksheet:
    for cell in ws[col_letter]:
        cell.number_format = fmt
    return ws


def write_df_to_excel(
    filename: Union[str, Path],
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    start_row: Optional[int] = None,
    start_col: Optional[int] = None,
    min_row: Optional[int] = None,
    add_filter: bool = False,
    fmt_date: str = "dd-mm-yyyy",
    fmt_datetime: str = "dd-mm-yyyy hh:mm",
    clear_sheet_values: bool = False,
    **to_excel_kwargs,
) -> None:
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet. If [filename] doesn't exist, then this function will create it.

    storage_options: dict, optional
        Extra options that make sense for a particular storage connection, e.g. host, port,
        username, password, etc., if using a URL that will be parsed by fsspec, e.g.,
        starting “s3://”, “gcs://”.
    to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]

    Usage examples:

    >>> append_df_to_excel('/tmp/test.xlsx', df, add_filter=True,
                           freeze_panes=(1,0))

    >>> append_df_to_excel('/tmp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, start_row=25)

    >>> append_df_to_excel('/tmp/test.xlsx', df, index=False,
                           fmt_datetime="dd.mm.yyyy hh:mm")

    """
    file_exists = check_if_file_exists(filename)
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = get_sheets_in_workbook(wb)

    with pd.ExcelWriter(
        filename,
        # filename.with_suffix(".xlsx"),
        mode="a" if file_exists else "w",
        if_sheet_exists="new" if file_exists else None,
        date_format=fmt_date,
        datetime_format=fmt_datetime,
    ) as writer:
        if file_exists:
            writer.book = wb

            if sheet_exists:
                if not start_row:
                    start_row = get_max_row(ws=writer.book[sheet_name]) + 1

                if not start_col:
                    start_col = get_start_col(ws=writer.book[sheet_name])

                if clear_sheet_values:
                    writer.book = clear_sheet(wb=writer.book, sheet_name=sheet_name)
            writer.sheets = sheets
        else:
            # file doesn't exist -> create new one
            start_row, start_col = 1, 1

        # write out the dataframe to excelwriter
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        worksheet = writer.sheets[sheet_name]

        if add_filter:
            worksheet.auto_filter.ref = worksheet.dimensions

        first_col = int(to_excel_kwargs.get("index", True)) + 1
        worksheet = format_data_in_sheet(
            ws=worksheet, df=df, index=to_excel_kwargs.get("index", True)
        )

    # TO DO: fix append part of the function
    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)
        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]
        # copy rows written by `df.to_excel(...)` to
        _copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
            tgt_min_row=start_row,
            with_style=True,
            min_row=min_row,
        )
        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()


if __name__ == "__main__":

    df = pd.DataFrame([[1000, 20], [30, 40]])
    filename = "tim.xlsx"
    sheet_name = "Grades"
    to_excel_kwargs = {"index": False}
    # to_excel_kwargs = {}
    write_df_to_excel(filename, df, sheet_name, **to_excel_kwargs)
